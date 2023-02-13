from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import load_workbook
import json
import requests

def index(request):
    if request.method == 'POST':
        api_key = request.POST.get('api_key')
        xlsx_file = request.FILES['file']
        wb = load_workbook(filename=xlsx_file, read_only=True)
        ws = wb.active
        results = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_list = {'document': row[0]}
            inputs_list = {'data': data_list}
            req_list = {'Inputs': inputs_list, 'GlobalParameters': 1.0}
            req_json = json.dumps(req_list)
            headers = {'Content-Type': 'application/json', 'Authorization': f'Bearer {api_key}'}
            response = requests.post('http://85703749-f0d5-489f-8a7f-724648002ab4.koreacentral.azurecontainer.io/score', json=req_json, headers=headers)
            if response.status_code >= 400:
                return HttpResponse(f'The request failed with status code: {response.status_code}')
            result = response.json()['Results']['output1']['value']['Values'][0][0]
            results.append(result)
        wb['A1'] = 'Result'
        for i, result in enumerate(results):
            wb.cell(row=i+2, column=1, value=result)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=result.xlsx'
        wb.save(response)
        return response
    return render(request, 'index.html')
